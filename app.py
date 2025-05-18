import pandas as pd
import openpyxl
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import time
import logging
from datetime import datetime
import os
from flask import Flask, render_template, request, send_file, session, redirect, url_for
from io import BytesIO
import html
from werkzeug.middleware.proxy_fix import ProxyFix

# Configure application
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('website_scan.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Constants
MAX_SCAN_TIME = 30  # seconds
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
REQUEST_TIMEOUT = 20
MAX_REDIRECTS = 5

# Enhanced scoring categories
CATEGORIES = {
    "First Impressions & Branding": {
        "scores": [
            "Lacks professional design and messaging. Recommend full redesign.",
            "Unclear offer. Suggest branding update, clearer value prop, and trust elements.",
            "Decent design, but needs polish. Recommend refining layout and visuals.",
            "Strong branding with minor design updates suggested.",
            "Excellent branding. No improvements needed."
        ],
        "weight": 1.2
    },
    "User Experience (UX)": {
        "scores": [
            "Confusing journey. Full UX overhaul needed.",
            "Navigation and flow inconsistent. Suggest restructuring.",
            "Usable, but some friction. Recommend usability testing.",
            "Good UX with minor friction points. Suggest tweaks.",
            "Excellent UX. No changes needed."
        ],
        "weight": 1.3
    },
    "Performance & Speed": [
        "Extremely slow. Recommend full optimization (hosting, images, scripts).",
        "Slow load times. Suggest compressing assets and optimizing code.",
        "Acceptable speed. Room for improvement with lazy loading/CDN.",
        "Good performance with small issues to address.",
        "Excellent performance. No changes needed."
    ],
    "Mobile Responsiveness": [
        "Poor experience on mobile. Recommend responsive redesign.",
        "Major mobile issues. Redesign mobile layout and fix touch elements.",
        "Responsive but with usability gaps. Suggest mobile-specific adjustments.",
        "Good responsiveness. Test and refine further.",
        "Perfect mobile design. No improvements needed."
    ],
    "SEO & Visibility": [
        "No SEO foundations. Recommend full setup (meta, sitemap, schema).",
        "Minimal SEO. Recommend on-page SEO and metadata improvements.",
        "Basic SEO setup. Recommend keyword and content optimization.",
        "Good SEO. Suggest content strategy enhancements.",
        "Excellent SEO. No improvements needed."
    ],
    "Security & Compliance": [
        "No HTTPS or compliance. Urgent fixes needed (SSL, policy, updates).",
        "Basic security but missing compliance features. Recommend updates.",
        "Secure but outdated components. Suggest plugin/CMS updates.",
        "Secure with minor improvements needed.",
        "Fully secure and compliant. No changes needed."
    ],
    "Accessibility": [
        "No accessibility. Recommend WCAG audit and full compliance plan.",
        "Major issues (contrast, keyboard nav). Recommend improvements.",
        "Some basics present. Suggest screen reader and contrast review.",
        "Mostly compliant. Suggest accessibility testing tools.",
        "Fully compliant and accessible. Great work!"
    ],
    "Analytics & Conversions": [
        "No tracking. Recommend GA4, goal setup, CRM integration.",
        "Basic analytics only. Add events and conversion goals.",
        "Some tracking in place. Recommend UTM and funnel tracking.",
        "Well-tracked site. Suggest dashboards and heatmaps.",
        "Excellent analytics. Fully optimized."
    ]
}

def initialize_results():
    """Create a fully initialized results dictionary with all required keys"""
    return {
        'basic': {
            'load_time': 0.0,
            'scan_timestamp': datetime.now().isoformat()
        },
        'meta': {
            'title': '',
            'title_length': 0,
            'description': '',
            'viewport': False,
            'has_favicon': False,
            'canonical': '',
            'og_tags': {}
        },
        'resources': {
            'images': 0,
            'stylesheets': 0,
            'scripts': 0
        },
        'performance': {
            'page_size_kb': 0.0,
            'requests': 0,
            'dom_elements': 0,
            'dom_depth': 0
        },
        'security': {
            'https': False,
            'hsts': False,
            'content_security_policy': False,
            'x_frame_options': False
        },
        'accessibility': {
            'alt_text_images': 0,
            'lang_attribute': False,
            'aria_attributes': 0
        },
        'issues': []
    }

def validate_url(url):
    """Robust URL validation with sanitization"""
    if not url or not isinstance(url, str):
        raise ValueError("URL cannot be empty")

    url = url.strip()
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url

    try:
        result = urlparse(url)
        if not all([result.scheme, result.netloc]):
            raise ValueError("Invalid URL structure")

        if '.' not in result.netloc or len(result.netloc) < 4:
            raise ValueError("Invalid domain format")

        return html.escape(url)
    except Exception as e:
        logger.error(f"URL validation failed: {str(e)}")
        raise ValueError(f"Invalid URL: {str(e)}")

def safe_html_parse(html_content):
    """Robust HTML parsing with multiple fallbacks"""
    parsers = ['lxml', 'html.parser', 'html5lib']
    last_error = None

    for parser in parsers:
        try:
            soup = BeautifulSoup(html_content, parser)
            if soup.find():
                return soup
        except Exception as e:
            last_error = e
            continue

    logger.warning(f"All parsers failed, returning empty soup. Last error: {str(last_error)}")
    return BeautifulSoup("", 'html.parser')

def scan_website(url):
    """Perform comprehensive website scan with complete error handling"""
    results = initialize_results()

    required_keys = {
        'performance': {'page_size_kb': 0.0, 'requests': 0, 'dom_depth': 0},
        'basic': {'load_time': 0.0},
    }
    for section, defaults in required_keys.items():
        results.setdefault(section, {})
        for key, val in defaults.items():
            if not isinstance(results[section].get(key), (int, float)):
                results[section][key] = val

    try:
        validated_url = validate_url(url)
        headers = {'User-Agent': USER_AGENT}

        start_time = time.time()
        logger.info(f"Starting scan of: {validated_url}")

        try:
            with requests.Session() as session:
                session.max_redirects = MAX_REDIRECTS
                response = session.get(
                    validated_url,
                    headers=headers,
                    timeout=REQUEST_TIMEOUT,
                    allow_redirects=True,
                    stream=True
                )
                response.raise_for_status()

                # Always capture these basic metrics
                results['basic']['load_time'] = time.time() - start_time
                content = response.content
                results['performance']['page_size_kb'] = len(content) / 1024
                results['security']['https'] = response.url.startswith('https://')

                try:
                    soup = safe_html_parse(content.decode('utf-8', errors='replace'))

                    # Meta data extraction with safe defaults
                    if soup.title:
                        results['meta']['title'] = soup.title.string or ''
                        results['meta']['title_length'] = len(results['meta']['title'])

                    # Other meta tags with safe access
                    meta_desc = soup.find('meta', attrs={'name': 'description'})
                    results['meta']['description'] = meta_desc.get('content', '') if meta_desc else ''

                    results['meta']['viewport'] = bool(soup.find('meta', {'name': 'viewport'}))
                    results['meta']['has_favicon'] = bool(soup.find('link', rel='icon'))

                    canonical = soup.find('link', rel='canonical')
                    results['meta']['canonical'] = canonical.get('href', '') if canonical else ''

                    # Accessibility checks with safe defaults
                    results['accessibility']['alt_text_images'] = len(soup.find_all('img', alt=True))

                    html_tag = soup.find('html')
                    results['accessibility']['lang_attribute'] = bool(html_tag.get('lang', '')) if html_tag else False
                    results['accessibility']['aria_attributes'] = len(soup.find_all(lambda tag: any(attr.startswith('aria-') for attr in tag.attrs)))

                    # Resources counting
                    results['resources']['images'] = len(soup.find_all('img'))
                    results['resources']['stylesheets'] = len(soup.find_all('link', rel='stylesheet'))
                    results['resources']['scripts'] = len(soup.find_all('script', src=True))

                    # Security headers
                    headers = response.headers
                    results['security']['hsts'] = 'strict-transport-security' in headers
                    results['security']['content_security_policy'] = 'content-security-policy' in headers
                    results['security']['x_frame_options'] = 'x-frame-options' in headers

                except Exception as parse_error:
                    logger.error(f"HTML parsing error: {str(parse_error)}")
                    results['issues'].append(f"HTML parsing error: {str(parse_error)}")
                    # Continue with partial results

        except requests.exceptions.RequestException as req_error:
            logger.error(f"Request failed: {str(req_error)}")
            results['issues'].append(f"Request failed: {str(req_error)}")
            results['basic']['load_time'] = time.time() - start_time
            results['performance']['page_size_kb'] = 0.0

        # Final validation before return
        perf = results.setdefault('performance', {})
        page_size = perf.get('page_size_kb', 0.0)
        if not isinstance(page_size, (int, float)):
            logger.warning("Invalid page_size_kb detected or missing, resetting to 0.0")
            perf['page_size_kb'] = 0.0

        return {'status': 'success', 'data': results}

    except Exception as e:
        logger.error(f"Scan error: {str(e)}", exc_info=True)
        results['issues'].append(f"Scan error: {str(e)}")
        results['performance']['page_size_kb'] = 0.0
        return {'status': 'error', 'message': str(e), 'data': results}

def auto_score_website(analysis_data, response_text=None):
    """Completely safe scoring with comprehensive validation"""
    # Initialize default structure if invalid input
    if not isinstance(analysis_data, dict):
        logger.warning("Invalid analysis_data - initializing default structure")
        analysis_data = initialize_results()

    # Ensure all sections exist with proper defaults
    required_sections = {
        'basic': {'load_time': 8.0},
        'performance': {
            'page_size_kb': 1000.0,
            'requests': 30,
            'dom_depth': 20
        },
        'security': {},
        'meta': {},
        'accessibility': {},
        'resources': {}
    }

    for section, defaults in required_sections.items():
        analysis_data.setdefault(section, {})
        for key, default_value in defaults.items():
            # Special handling for page_size_kb
            if key == 'page_size_kb':
                current_val = analysis_data[section].get(key)
                if not isinstance(current_val, (int, float)):
                    logger.warning(f"Invalid page_size_kb, resetting to default")
                    analysis_data[section][key] = float(default_value)
            else:
                analysis_data[section].setdefault(key, default_value)

    # Create local variables with safe values
    basic = analysis_data['basic']
    perf = analysis_data['performance']
    security = analysis_data['security']
    meta = analysis_data['meta']
    accessibility = analysis_data['accessibility']
    resources = analysis_data['resources']

    scores = {}

    # Performance scoring (now completely safe)
    load_time = basic['load_time']
    page_size = perf['page_size_kb']
    requests = perf['requests']
    dom_depth = perf['dom_depth']

    # Calculate performance score
    perf_score = 0

    # Load time scoring
    if load_time < 1.5:
        perf_score += 4
    elif load_time < 3:
        perf_score += 3
    elif load_time < 5:
        perf_score += 2
    elif load_time < 8:
        perf_score += 1

    # Page size scoring
    if page_size < 300:
        perf_score += 2
    elif page_size < 800:
        perf_score += 1

    # Requests scoring
    if requests < 15:
        perf_score += 2
    elif requests < 30:
        perf_score += 1

    # DOM depth scoring
    if dom_depth < 15:
        perf_score += 1
    elif dom_depth > 30:
        perf_score -= 1

    scores['Performance & Speed'] = min(max(1, perf_score // 2), 5)

    # Security scoring
    sec_score = 1  # Base score
    sec_score += 3 if security.get('https', False) else 0

    # Other security headers
    sec_score += 1 if security.get('hsts', False) else 0
    sec_score += 1 if security.get('content_security_policy', False) else 0
    sec_score += 1 if security.get('x_frame_options', False) else 0

    scores['Security & Compliance'] = min(sec_score, 5)

    # SEO scoring
    seo_score = 1  # Base score

    # Title check
    title = meta.get('title', '')
    if title and 30 <= len(title) <= 60:
        seo_score += 2

    # Description check
    description = meta.get('description', '')
    if description and 50 <= len(description) <= 160:
        seo_score += 2

    # Viewport is critical for mobile SEO
    if meta.get('viewport', False):
        seo_score += 1

    # Canonical URL
    if meta.get('canonical', ''):
        seo_score += 1

    # OpenGraph tags
    if meta.get('og_tags', {}):
        seo_score += 1

    scores['SEO & Visibility'] = min(seo_score, 5)

    # Mobile responsiveness
    mobile_score = 1
    if meta.get('viewport', False):
        mobile_score += 3

    # Check for responsive design indicators
    if resources.get('stylesheets', 0) > 0:
        mobile_score += 1

    scores['Mobile Responsiveness'] = min(mobile_score, 5)

    # First impressions
    first_imp_score = 2  # Base score

    if meta.get('has_favicon', False):
        first_imp_score += 1
    if title:
        first_imp_score += 1
    if description:
        first_imp_score += 1

    scores['First Impressions & Branding'] = min(first_imp_score, 5)

    # Content quality (basic assessment)
    content_score = 3  # Default

    # Check for meaningful content (only if response_text is provided)
    if response_text:
        try:
            soup = BeautifulSoup(response_text, 'lxml')
            text_content = ' '.join([p.get_text() for p in soup.find_all(['p', 'h1', 'h2', 'h3'])])
            word_count = len(text_content.split())

            if word_count > 500:
                content_score += 1
            elif word_count < 100:
                content_score -= 1
        except:
            pass

    scores['Content Quality'] = min(max(1, content_score), 5)

    # Accessibility
    accessibility_score = 1  # Base score

    # Image alt text
    total_images = resources.get('images', 0)
    alt_images = accessibility.get('alt_text_images', 0)
    if total_images > 0:
        alt_ratio = alt_images / total_images
        if alt_ratio > 0.9:
            accessibility_score += 2
        elif alt_ratio > 0.5:
            accessibility_score += 1

    # Language attribute
    if accessibility.get('lang_attribute', False):
        accessibility_score += 1

    # ARIA attributes
    if accessibility.get('aria_attributes', 0) > 0:
        accessibility_score += 1

    scores['Accessibility'] = min(accessibility_score, 5)

    return scores

def create_results_dataframe(scores, url):
    """Create DataFrame from scoring results"""
    data = []
    for category, score in scores.items():
        if isinstance(CATEGORIES.get(category, {}), dict):
            recommendation = CATEGORIES[category]['scores'][score-1]
        else:
            recommendation = CATEGORIES.get(category, ["No recommendation available"])[score-1]

        data.append({
            'Section': category,
            'Score': score,
            'Recommendation': recommendation
        })

    return pd.DataFrame(data)

def create_styled_spreadsheet(df, scan_data, url):
    """Create professionally styled Excel workbook"""
    wb = openpyxl.Workbook()

    # ===== Scorecard Sheet =====
    ws = wb.active
    ws.title = "Scorecard"

    # Define styles
    header_fill = PatternFill(start_color="2A5CAA", end_color="2A5CAA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Write headers
    headers = ["Category", "Score", "Recommendation", "Priority", "Details"]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Conditional formatting for scores
    score_colors = {
        1: "FF0000",  # Red
        2: "FF6600",  # Orange
        3: "FFCC00",  # Yellow
        4: "92D050",  # Light green
        5: "00B050"   # Dark green
    }

    # Write data with enhanced formatting
    for idx, row in df.iterrows():
        ws.append([
            row['Section'],
            row['Score'],
            row['Recommendation'],
            "High" if row['Score'] in [1, 2] else "Medium" if row['Score'] == 3 else "Low",
            get_category_details(row['Section'], scan_data)
        ])

        # Style all cells in row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=idx+2, column=col)
            cell.font = Font(name='Calibri', size=11)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if row['Score'] in [1, 2]:
                cell.fill = highlight_fill

        # Special score cell styling
        score_cell = ws.cell(row=idx+2, column=2)
        score_cell.fill = PatternFill(
            start_color=score_colors.get(row['Score'], "FFFFFF"),
            end_color=score_colors.get(row['Score'], "FFFFFF"),
            fill_type="solid"
        )
        score_cell.font = Font(bold=True)
        score_cell.alignment = Alignment(horizontal="center")

    # Set column widths and row heights
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40

    for row in range(2, len(df) + 2):
        ws.row_dimensions[row].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ===== Scan Details Sheet =====
    ws2 = wb.create_sheet(title="Scan Details")

    sections = [
        ("Basic Information", "basic"),
        ("Meta Tags Analysis", "meta"),
        ("Resources Breakdown", "resources"),
        ("Performance Metrics", "performance"),
        ("Security Headers", "security"),
        ("Accessibility Checks", "accessibility")
    ]

    row_num = 1
    for section_name, section_key in sections:
        # Section header
        ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        header_cell = ws2.cell(row=row_num, column=1, value=section_name)
        header_cell.font = Font(bold=True, size=14, color="2A5CAA")
        header_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row_num += 1

        # Sub-headers
        ws2.cell(row=row_num, column=1, value="Metric").font = Font(bold=True)
        ws2.cell(row=row_num, column=2, value="Value").font = Font(bold=True)
        row_num += 1

        # Data rows
        if section_key in scan_data:
            for key, value in scan_data[section_key].items():
                key_cell = ws2.cell(row=row_num, column=1, value=key.replace('_', ' ').title())
                key_cell.font = Font(bold=False)

                value_cell = ws2.cell(row=row_num, column=2, value=str(value))

                if isinstance(value, str) and value.startswith(('http://', 'https://')):
                    value_cell.hyperlink = value
                    value_cell.font = Font(color="0563C1", underline="single")
                elif isinstance(value, bool):
                    value_cell.font = Font(color="00B050" if value else "FF0000")
                    value_cell.value = "Yes" if value else "No"

                row_num += 1

        row_num += 1

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 50
    ws2.freeze_panes = "A2"

    # ===== Executive Summary Sheet =====
    ws3 = wb.create_sheet(title="Summary")
    ws3.merge_cells(start_row=1, start_column=1, end_row=3, end_column=2)
    ws3.cell(row=1, column=1, value="Website Review Report").font = Font(bold=True, size=18, color="2A5CAA")

    # Add summary metrics
    summary_metrics = [
        ["Website URL", url],
        ["Scan Date", scan_data.get('basic', {}).get('scan_timestamp', 'N/A')],
        ["Overall Score", f"{df['Score'].mean():.1f}/5.0"],
        ["Page Title", scan_data.get('meta', {}).get('title', 'N/A')],
        ["Load Time", f"{scan_data.get('basic', {}).get('load_time', 0):.2f} seconds"],
        ["Page Size", f"{scan_data.get('performance', {}).get('page_size_kb', 0):.1f} KB"],
        ["Resource Requests", scan_data.get('performance', {}).get('requests', 0)],
        ["Uses HTTPS", "Yes" if scan_data.get('security', {}).get('https', False) else "No (Critical)"],
        ["Mobile Ready", "Yes" if scan_data.get('meta', {}).get('viewport', False) else "No"],
        ["Critical Issues", len(df[df['Score'] == 1])],
        ["Areas Needing Improvement", len(df[df['Score'] == 2])],
        ["Well Performing Areas", len(df[df['Score'] >= 4])]
    ]

    start_row = 5
    for i, (metric, value) in enumerate(summary_metrics, start=start_row):
        ws3.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=value)

        try:
            if metric == "Uses HTTPS" and str(value).startswith("No"):
                ws3.cell(row=i, column=2).font = Font(color="FF0000", bold=True)
        except Exception as e:
            logger.warning(f"Font styling failed at summary row {i}: {e}")


    # Add recommendations section
    rec_row = start_row + len(summary_metrics) + 2
    ws3.cell(row=rec_row, column=1, value="Top Recommendations").font = Font(bold=True, size=14, color="2A5CAA")
    rec_row += 1

    top_issues = df.nsmallest(3, 'Score')
    for idx, row in top_issues.iterrows():
        ws3.cell(row=rec_row, column=1, value=f"• {row['Recommendation']}").font = Font(bold=True)
        ws3.cell(row=rec_row, column=1).alignment = Alignment(wrap_text=True)
        rec_row += 1

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 40

    return wb

def get_category_details(category, scan_data):
    """Completely safe category details generator with nested .get()"""
    details = []

    try:
        # Safely access nested dictionaries with .get()
        basic = scan_data.get('basic', {})
        performance = scan_data.get('performance', {})
        security = scan_data.get('security', {})
        meta = scan_data.get('meta', {})
        resources = scan_data.get('resources', {})
        accessibility = scan_data.get('accessibility', {})

        if category == "Performance & Speed":
            details.extend([
                f"Load Time: {basic.get('load_time', 0):.2f}s",
                f"Page Size: {performance.get('page_size_kb', 0):.1f}KB",
                f"Requests: {performance.get('requests', 0)}",
                f"DOM Depth: {performance.get('dom_depth', 0)}"
            ])

        elif category == "Security & Compliance":
            details.append(f"HTTPS: {'Yes' if security.get('https', False) else 'No'}")
            if security.get('https', False):
                details.append(f"HSTS: {'Yes' if security.get('hsts', False) else 'No'}")
            details.append(f"CSP Header: {'Yes' if security.get('content_security_policy', False) else 'No'}")

        elif category == "Mobile Responsiveness":
            details.extend([
                f"Viewport: {'Present' if meta.get('viewport', False) else 'Missing'}",
                f"Images: {resources.get('images', 0)}",
                f"Responsive CSS: {resources.get('stylesheets', 0)} sheets"
            ])

        elif category == "First Impressions & Branding":
            details.extend([
                f"Title: {'Present' if meta.get('title', '') else 'Missing'}",
                f"Favicon: {'Present' if meta.get('has_favicon', False) else 'Missing'}"
            ])

        elif category == "Accessibility":
            total_images = resources.get('images', 0)
            alt_images = accessibility.get('alt_text_images', 0)
            details.extend([
                f"Alt Text: {alt_images}/{total_images} images",
                f"ARIA Attributes: {accessibility.get('aria_attributes', 0)}",
                f"Language: {'Set' if accessibility.get('lang_attribute', False) else 'Missing'}"
            ])

        elif category == "SEO & Visibility":
            details.extend([
                f"Title Length: {meta.get('title_length', 0)} chars",
                f"Description: {'Present' if meta.get('description', '') else 'Missing'}",
                f"Viewport: {'Present' if meta.get('viewport', False) else 'Missing'}"
            ])

    except Exception as e:
        logger.error(f"Error generating details for {category}: {str(e)}")
        details.append("Details currently unavailable")

    return "\n".join(details)

def ensure_defaults(scan_data):
    """Ensures scan_data contains valid values with correct types for all required fields"""
    defaults = {
        'basic': {'load_time': 0.0},
        'performance': {'page_size_kb': 0.0, 'requests': 0, 'dom_depth': 0},
        'meta': {'viewport': False, 'title': '', 'description': ''},
        'resources': {'stylesheets': 0},
        'security': {'https': False},
        'accessibility': {'alt_text_images': 0, 'aria_attributes': 0, 'lang_attribute': False}
    }
    for section, fields in defaults.items():
        sec = scan_data.setdefault(section, {})
        for key, val in fields.items():
            # Force override if missing or wrong type
            if key not in sec or not isinstance(sec[key], type(val)):
                logger.warning(f"Fixing missing or invalid key: {section}.{key} = {sec.get(key)}")
                sec[key] = val
    logger.info(f"Sanitized scan_data: {scan_data}")
    return scan_data

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        url = request.form.get('url', '').strip()
        if not url:
            return render_template('index.html', error="Please enter a URL")

        try:
            validated_url = validate_url(url)
            scan_results = scan_website(validated_url)

            logger.info(f"Full scan results: {scan_results}")
            if 'data' not in scan_results:
                raise ValueError("Invalid scan results format - missing data")

            scan_data = scan_results.get('data', {})
            scan_data = ensure_defaults(scan_data)

            logger.info(f"Post-default performance keys: {scan_data.get('performance', {})}")

            if 'page_size_kb' not in scan_data.get('performance', {}):
                logger.error("page_size_kb STILL missing in performance")
                scan_data['performance']['page_size_kb'] = 0.0

            scores = auto_score_website(scan_data)
            df = create_results_dataframe(scores, validated_url)

            session['scan_results'] = {
                'df': df.to_dict(),
                'scan_data': scan_data,
                'validated_url': validated_url
            }

            chart_data = {
                'labels': list(scores.keys()),
                'data': list(scores.values()),
                'colors': ['#e74c3c', '#e67e22', '#f1c40f', '#2ecc71', '#27ae60']
            }

            return render_template('results.html',
                                   url=validated_url,
                                   scores=scores,
                                   chart_data=chart_data,
                                   scan_data=scan_data)

        except Exception as e:
            logger.error(f"Scan failed: {str(e)}", exc_info=True)
            return render_template('index.html', error=str(e))

    return render_template('index.html')

@app.route('/download')
def download_report():
    if 'scan_results' not in session:
        return redirect(url_for('index'))

    try:
        # Reconstruct DataFrame from session
        df = pd.DataFrame(session['scan_results']['df'])
        scan_data = session['scan_results']['scan_data']
        url = session['scan_results']['validated_url']

        # Create Excel report in memory
        output = BytesIO()
        wb = create_styled_spreadsheet(df, scan_data, url)
        wb.save(output)
        output.seek(0)

        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Website_Review_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Download failed: {str(e)}", exc_info=True)
        return redirect(url_for('index', error="Failed to generate report"))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('DEBUG', False))